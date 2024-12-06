from typing import List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO
from aiogram import Router, F, Bot, types
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message, FSInputFile, \
    BufferedInputFile
from aiogram.filters import Command, CommandStart, StateFilter
from lexicon.lexicon_ru import LEXICON_RU
from keyboards.inline_keyboards import create_inline_kb
from database.database import select_drivers, add_user, get_users, send_predict, get_predict, add_result, \
    show_result, get_actual_gp, add_points, show_result, show_points, get_result, check_res, show_points_all, \
    is_prediced, get_user_team, add_team, get_team, show_points_team_all, get_teams_fonts_colors, clear_results, \
    get_name_gp, get_maximus
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import default_state, State, StatesGroup
from aiogram.fsm.storage.redis import RedisStorage, Redis
from dataprocessing.calculation_gp_drivers import calculation_drivers
from string import ascii_letters, digits

router: Router = Router()

# Инициализируем Redis
redis = Redis(host='127.0.0.1')

# Инициализируем хранилище (создаем экземпляр класса MemoryStorage)
storage = RedisStorage(redis=redis)


class FSMFillForm(StatesGroup):
    # Создаем экземпляры класса State, последовательно
    # перечисляя возможные состояния, в которых будет находиться
    # бот в разные моменты взаимодейтсвия с пользователем
    fill_name = State()  # Состояние ожидания ввода имени
    fill_second_name = State()  # Состояние ожидания ввода фамилии
    select_first = State()
    select_second = State()
    select_third = State()
    select_fourth = State()
    select_team = State()
    select_engine = State()
    select_gap = State()
    select_lapped = State()
    end_select = State()
    fill_team_name = State()
    fill_team_number = State()


# Этот хэндлер срабатывает на команду /start
@router.message(CommandStart(), StateFilter(default_state))
async def process_start_command(message: Message):
    await message.answer(text=LEXICON_RU['start_answer'], reply_markup=types.ReplyKeyboardRemove())


# Этот хэндлер будет срабатывать на команду "/cancel" в любых состояниях,
# кроме состояния по умолчанию, и отключать машину состояний
@router.message(Command(commands='cancel'), ~StateFilter(default_state))
async def process_cancel_command_state(message: Message, state: FSMContext):
    await message.answer(
        text='Вы вышли из ввода данных\n\n'
             'Чтобы снова перейти к заполнению -  '
             'отправьте соответствующую команду'
    )
    # Сбрасываем состояние и очищаем данные, полученные внутри состояний
    await state.clear()


# Этот хэндлер срабатывает на команду /help
@router.message(Command(commands=['help']))
async def process_help_command(message: Message):
    await message.answer(text=LEXICON_RU['help_answer'])


# Этот хэндлер срабатывает на команду /clear_result
@router.message(Command(commands=['clear_result']))
async def clear_result_command(message: Message):
    clear_results(get_actual_gp())
    await message.answer('Результат удалён')


'''

///... ХЭНДЛЕРЫ РЕГИСТРАЦИИ НАЧАЛО ///

'''


# Этот хэндлер будет срабатывать на команду /registration
# и переводить бота в состояние ожидания ввода имени
@router.message(Command(commands='registration'), StateFilter(default_state))
async def process_fillform_command(message: Message, state: FSMContext):
    if not get_users(message.from_user.id):
        await message.answer(text='Пожалуйста, введите ваше имя латинским буквами')
        # Устанавливаем состояние ожидания ввода имени
        await state.set_state(FSMFillForm.fill_name)
    else:
        user = get_users(message.from_user.id)
        await message.answer(text=f'Вы уже зарегистрированы как {user.name}')


# Этот хэндлер будет срабатывать, если введено корректное имя
@router.message(StateFilter(FSMFillForm.fill_name), lambda message: all(char in ascii_letters for char in message.text))
async def process_lastname_sent(message: Message, state: FSMContext):
    # Сохраняем введенное имя в хранилище по ключу "name"
    await state.update_data(name=message.text.capitalize())
    await message.answer(text='Спасибо!\n\nА теперь введите вашу фамилию латинским буквами')
    await state.set_state(FSMFillForm.fill_second_name)


# Этот хэндлер будет срабатывать, если во время ввода имени
# будет введено что-то некорректное
@router.message(StateFilter(FSMFillForm.fill_name))
async def warning_not_name(message: Message):
    await message.answer(
        text='То, что вы отправили не похоже на имя латинским буквами\n\n'
             'Пожалуйста, введите ваше имя\n\n'
             'Если вы хотите прервать заполнение анкеты - '
             'отправьте команду /cancel')


# Этот хэндлер будет срабатывать на ввод фамилии
# записывать данные и выводить из машины состояний
@router.message(StateFilter(FSMFillForm.fill_second_name),
                lambda message: all(char in ascii_letters for char in message.text))
async def process_wish_news_press(message: Message, state: FSMContext):
    # Cохраняем данные о вк
    await state.update_data(lastname=message.text.upper())
    # Добавляем в базу данных анкету пользователя
    # по ключу id пользователя
    user = await state.get_data()
    add_user(message.from_user.id, **user)

    # Завершаем машину состояний
    await state.clear()
    # Отправляем в чат сообщение о сохранении данных
    await message.answer(
        text='Спасибо! Ваши данные сохранены!\n\n'
    )
    # Отправляем в чат сообщение с предложением посмотреть свою анкету
    await message.answer(
        text='Чтобы посмотреть данные вашей '
             'анкеты - отправьте команду /showdata'
    )


# Этот хэндлер будет срабатывать, если во время ввода фамилии
# будет введено что-то некорректное
@router.message(StateFilter(FSMFillForm.fill_second_name))
async def warning_not_name(message: Message):
    await message.answer(
        text='То, что вы отправили не похоже на фамилию латинским буквами\n\n'
             'Пожалуйста, введите вашу фамилию\n\n'
             'Если вы хотите прервать заполнение анкеты - '
             'отправьте команду /cancel')


'''

///... ХЭНДЛЕРЫ РЕГИСТРАЦИИ КОНЕЦ ///

'''

'''

///... ХЭНДЛЕРЫ РЕГИСТРАЦИИ КОМАНДЫ НАЧАЛО ///

'''


# Этот хэндлер будет срабатывать на команду /createteam
@router.message(Command(commands='createteam'), StateFilter(default_state))
async def process_createteam_command(message: Message, state: FSMContext):
    if not get_users(message.from_user.id):
        await message.answer(text='Вы не зарегистрированы, зарегистрируйтесь перед созданием команды.')
    else:
        if get_user_team(message.from_user.id) == 'PERSONAL ENTRY':
            await message.answer(text='Пожалуйста, введите название команды латинским буквами')
            await state.set_state(FSMFillForm.fill_team_name)
        else:
            await message.answer(text=f'Вы уже находитесь в команде <b>{get_user_team(message.from_user.id)}</b>')


# Этот хэндлер будет срабатывать, если введено корректное название
@router.message(StateFilter(FSMFillForm.fill_team_name),
                lambda message: all(char in ascii_letters + digits + "'" for char in message.text))
async def process_lastname_sent(message: Message, state: FSMContext):
    # Сохраняем введенное имя в хранилище по ключу "name"
    await state.update_data(name=message.text)
    await message.answer(text='Спасибо!\nВведите номер гонщика')
    await state.set_state(FSMFillForm.fill_team_number)


# Этот хэндлер будет срабатывать, если во время ввода имени
@router.message(StateFilter(FSMFillForm.fill_team_name))
async def warning_not_name(message: Message):
    await message.answer(
        text='То, что вы отправили не похоже на название команды латинским буквами\n\n'
             'Пожалуйста, введите название команды\n\n'
             'Если вы хотите прервать заполнение анкеты - '
             'отправьте команду /cancel')


# Этот хэндлер будет срабатывать на ввод фамилии
# записывать данные и выводить из машины состояний
@router.message(StateFilter(FSMFillForm.fill_team_number), F.text.isdigit())
async def process_wish_news_press(message: Message, state: FSMContext):
    # Cохраняем данные о вк
    await state.update_data(number=message.text.upper())
    # Добавляем в базу данных анкету пользователя
    # по ключу id пользователя
    user = await state.get_data()
    add_team(user_id=message.from_user.id, captain=True, **user)

    # Завершаем машину состояний
    await state.clear()
    # Отправляем в чат сообщение о сохранении данных
    await message.answer(
        text=f'Спасибо! Ваши команда <b>{get_user_team(message.from_user.id)}</b> зарегистрирована!\n\n'
    )


# Этот хэндлер будет срабатывать, если во время ввода фамилии
# будет введено что-то некорректное
@router.message(StateFilter(FSMFillForm.fill_second_name))
async def warning_not_name(message: Message):
    await message.answer(
        text='То, что вы отправили не похоже на фамилию латинским буквами\n\n'
             'Пожалуйста, введите вашу фамилию\n\n'
             'Если вы хотите прервать заполнение анкеты - '
             'отправьте команду /cancel')


'''
///... ХЭНДЛЕРЫ РЕГИСТРАЦИИ КОМАНДЫ КОНЕЦ ///
'''

'''
///... ДОБАВЛЕНИЕ ЧЛЕНА КОМАНДЫ ///

@router.message(Command(commands='add_teammate'), StateFilter(default_state))
async def process_add_teammate_command(message: Message, state: FSMContext):
    print(get_team(message.from_user.id).__dict__)
    if not get_users(message.from_user.id):
        await message.answer(text='Вы не зарегистрированы, зарегистрируйтесь для выполнения действия команды.')
    else:
        if get_user_team(message.from_user.id) != 'PERSONAL ENTRY' and get_team(message.from_user.id).captain \
            and not all([get_team(message.from_user.id).first, get_team(message.from_user.id).second, get_team(message.from_user.id).third]):
            await message.answer(text='Пожалуйста, введите имя гонщика, которого Вы хотите добавить')
            await state.set_state(FSMFillForm.fill_team_name)
        else:
            await message.answer(text=f'Вы не являетесь капитаном команды')
'''

'''
///... ХЭНДЛЕРЫ ПРОГНОЗА НАЧАЛО ///
'''


# Этот хэндлер срабатывает на команду /predict и начинаем собирать прогноз
# Проверка зарегистрирован ли пользователь и отправляем кнопки с Командами по алфавиту
@router.message(Command(commands=['predict']), StateFilter(default_state))
async def predict_team(message: Message, state: FSMContext):
    if get_users(message.from_user.id):
        if not is_prediced(message.from_user.id, get_actual_gp()):
            await message.answer(
                text='Выберите Команду',
                reply_markup=create_inline_kb(1, *sorted(
                    {i.driver_team + '  ' + i.engine_short for i in select_drivers()})))
            await state.set_state(FSMFillForm.select_engine)
        else:
            await message.answer(text='Вы уже отправили прогноз на актуальный GP')
    else:
        await message.answer(text='Вы не зарегистрированы')


# Сохраняем команду, отправляем кнопки с двигателями
@router.callback_query(StateFilter(FSMFillForm.select_engine),
                       F.data.in_({i.driver_team + '  ' + i.engine_short for i in select_drivers()}))
async def predict_engine(callback: CallbackQuery, state: FSMContext):
    await state.update_data(driver_team=callback.data.rsplit(' ', 1)[0].strip())
    await state.update_data(select1_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    await callback.message.answer(text='Спасибо!\nТеперь выберите двигатель',
                                  reply_markup=create_inline_kb(1,
                                                                *sorted(
                                                                    {i.driver_engine + '  ' + i.engine_short for i in
                                                                     select_drivers()})))
    await state.set_state(FSMFillForm.select_first)


# Сохраняем двигатель, отправляем кнопки с выбором первого пилота
@router.callback_query(StateFilter(FSMFillForm.select_first),
                       F.data.in_({i.driver_engine + '  ' + i.engine_short for i in select_drivers()}))
async def predict_first(callback: CallbackQuery, state: FSMContext):
    await state.update_data(driver_engine=callback.data.rsplit(' ', 1)[0].strip())
    await state.update_data(select2_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    await callback.message.answer(text='Спасибо!\nТеперь выберите первого пилота',
                                  reply_markup=create_inline_kb(1,
                                                                *[
                                                                    i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short
                                                                    for i in
                                                                    select_drivers()]))
    await state.set_state(FSMFillForm.select_second)


# Сохраняем первого пилота, отправляем кнопки с выбором второго пилота
@router.callback_query(StateFilter(FSMFillForm.select_second),
                       F.data.in_([i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short for i in
                                   select_drivers()]))
async def predict_second(callback: CallbackQuery, state: FSMContext):
    await state.update_data(first_driver=callback.data.split('(')[0].strip())
    await state.update_data(select3_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    predict = await state.get_data()
    await callback.message.answer(text='Спасибо!\nТеперь выберите второго пилота',
                                  reply_markup=create_inline_kb(1,
                                                                *[
                                                                    i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short
                                                                    for i in
                                                                    select_drivers() if
                                                                    i.driver_name not in [*predict.values()]]))
    await state.set_state(FSMFillForm.select_third)


# Сохраняем второго пилота, отправляем кнопки с выбором третьего пилота
@router.callback_query(StateFilter(FSMFillForm.select_third),
                       F.data.in_([i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short for i in
                                   select_drivers()]))
async def predict_third(callback: CallbackQuery, state: FSMContext):
    await state.update_data(second_driver=callback.data.split('(')[0].strip())
    await state.update_data(select4_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    predict = await state.get_data()
    if all(engine == predict['select1_engine'] for engine in
           [predict['select2_engine'], predict['select3_engine'], predict['select4_engine']]):
        await callback.message.answer(
            text='Вы выбрали 4 участника с одним мотором, начните выбор заново с команды /predict')
        await state.clear()
    else:
        await callback.message.answer(text='Спасибо!\nТеперь выберите третьего пилота',
                                      reply_markup=create_inline_kb(1,
                                                                    *[
                                                                        i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short
                                                                        for i
                                                                        in select_drivers()[10:] if
                                                                        i.driver_name not in [*predict.values()]]))
        await state.set_state(FSMFillForm.select_fourth)


# Сохраняем третьего пилота, отправляем кнопки с выбором четвертого пилота
@router.callback_query(StateFilter(FSMFillForm.select_fourth),
                       F.data.in_([i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short for i in
                                   select_drivers()][10:]))
async def predict_fourth(callback: CallbackQuery, state: FSMContext):
    await state.update_data(third_driver=callback.data.split('(')[0].strip())
    await state.update_data(select5_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    predict = await state.get_data()
    values = [predict['select1_engine'], predict['select2_engine'], predict['select3_engine'],
              predict['select4_engine'], predict['select5_engine']]
    if any(values.count(x) == 4 for x in set(values)):
        await callback.message.answer(
            text='Вы выбрали 4 участника с одним мотором, начните выбор заново с команды /predict')
        await state.clear()
    else:
        await callback.message.answer(text='Спасибо!\nТеперь выберите четвертого пилота',
                                      reply_markup=create_inline_kb(1,
                                                                    *[
                                                                        i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short
                                                                        for i
                                                                        in select_drivers()[15:] if
                                                                        i.driver_name not in [*predict.values()]]))
        await state.set_state(FSMFillForm.select_gap)


# Сохраняем четвертого пилота, отправляем текст с выбором отставания от лидера
@router.callback_query(StateFilter(FSMFillForm.select_gap),
                       F.data.in_([i.driver_name + ' (' + i.driver_team + ')' + '  ' + i.engine_short for i in
                                   select_drivers()][15:]))
async def predict_gap(callback: CallbackQuery, state: FSMContext):
    await state.update_data(fourth_driver=callback.data.split('(')[0].strip())
    await state.update_data(select6_engine=callback.data.split()[-1].strip())
    await callback.message.delete()
    predict = await state.get_data()
    values = [predict['select1_engine'], predict['select2_engine'], predict['select3_engine'],
              predict['select4_engine'], predict['select5_engine'], predict['select6_engine']]
    if any(values.count(x) == 4 for x in set(values)):
        await callback.message.answer(
            text='Вы выбрали 4 участника с одним мотором, начните выбор заново с команды /predict')
        await state.clear()
    else:
        await callback.message.answer(text='Спасибо!\nТеперь введите отставание от лидера')
        await state.set_state(FSMFillForm.select_lapped)


# Сохраняем отставание, отправляем текст с выбором количества круговых
@router.message(StateFilter(FSMFillForm.select_lapped), F.text.isdigit())
async def predict_gap(message: CallbackQuery, state: FSMContext):
    await state.update_data(gap=message.text)
    await message.answer(text='Спасибо!\nТеперь введите количество круговых')
    await state.set_state(FSMFillForm.end_select)


# Сохранение количества круговых и запись прогноза в БД, выход
@router.message(StateFilter(FSMFillForm.end_select), F.text.isdigit())
async def predict_lap(message: CallbackQuery, state: FSMContext):
    await state.update_data(lapped=message.text)
    await state.update_data(penalty=False)
    predict = await state.get_data()
    values = [predict['select1_engine'], predict['select2_engine'], predict['select3_engine'],
              predict['select4_engine'], predict['select5_engine'], predict['select6_engine']]
    if len({predict['first_driver'], predict['second_driver'], predict['third_driver'], predict['fourth_driver']}) != 4:
        await message.answer(
            text='<b>Вы выбрали несколько одинаковых гонщиков, используя старые кнопки, пожалуйста, используйте актуальные кнопки и начните выбор заново с команды /predict</b>')
        await state.clear()
    elif any(values.count(x) == 4 for x in set(values)):
        await message.answer(
            text='Вы выбрали 4 участника с одним мотором, начните выбор заново с команды /predict')
        await state.clear()
    else:
        await message.answer(text=f'''Спасибо!\nВы выбрали: 
        Команда: <b>{predict['driver_team']}</b> 
        Двигатель: <b>{predict['driver_engine']}</b>
        Первый пилот: <b>{predict['first_driver']}</b>
        Второй пилот: <b>{predict['second_driver']}</b>
        Третий пилот: <b>{predict['third_driver']}</b>
        Четвертый пилот: <b>{predict['fourth_driver']}</b>
        Отставание от лидера: <b>{predict['gap']}</b>
        Количество круговых: <b>{predict['lapped']}</b>
        ''')
        # Удаляем служебные ключи
        for i in ['select1_engine', 'select2_engine', 'select3_engine', 'select4_engine', 'select5_engine',
                  'select6_engine']:
            predict.pop(i)

        # Пишем прогноз в базу
        gp = get_actual_gp()
        send_predict(message.from_user.id, gp, **predict)

        # Завершаем машину состояний
        await state.clear()
        # Отправляем в чат сообщение с предложением посмотреть свою анкету
        await message.answer(
            text='Чтобы посмотреть свой прогноз '
                 ' - отправьте команду /viewpredict'
        )


'''

///ХЭНДЛЕРЫ ПРОГНОЗА КОНЕЦ///

'''


# Этот хэндлер будет срабатывать на отправку команды /showdata
# и отправлять в чат данные анкеты, либо сообщение об отсутствии данных
@router.message(Command(commands='showdata'), StateFilter(default_state))
async def process_showdata_command(message: Message):
    # Отправляем пользователю анкету, если она есть в "базе данных"
    if get_users(message.from_user.id):
        user = get_users(message.from_user.id)
        await message.answer(f' Ваше имя: {user.name}, Ваша команда: {user.user_team}')
    else:
        await message.answer(text='Вы не зарегистрированы')


# Этот хэндлер будет срабатывать на отправку команды /viewresult
@router.message(Command(commands='viewresult'), StateFilter(default_state))
async def process_viewresult_command(message: Message):
    gp = get_actual_gp()
    data = show_result(gp)

    text_for_answer = f'POS|DRIVER              |D1|D2|D3|D4|TM|EN|DF|LP|PN|PTS|CPT|\n'
    for index, (user, result, points) in enumerate(data, 1):
        text_for_answer += f'{index:<3}|{user.name:<20}|{result.first_driver:<2}|{result.second_driver:<2}|{result.third_driver:<2}|{result.fourth_driver:<2}|{result.driver_team:<2}|{result.driver_engine:<2}|{result.gap:<2}|{result.lapped:<2}|{result.penalty:<2}|{result.total:<3}|{points.points:<3}|\n'
    await message.answer(f'<code>{text_for_answer}</code>', isable_web_page_preview=True)


# Этот хэндлер будет срабатывать на отправку команды /resultxls
@router.message(Command(commands='resultxls'), StateFilter(default_state))
async def process_result_xls_command(message: Message):
    gp = get_actual_gp()
    data = show_result(gp)

    # Создаем новый Excel файл
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Results"

    # Записываем заголовки
    headers = ['POS', '№', 'DRIVER', 'TEAM', 'DR1', 'DR2', 'DR3', 'DR4', 'TM', 'ENG', 'DIFF', 'LAP', 'PEN', 'PTS',
               'CH.PTS']
    sheet.append(headers)

    # Записываем данные
    for index, (user, result, points) in enumerate(data, 1):
        sheet.append([
            index,
            user.number,
            user.name,
            get_user_team(user.id_telegram),
            result.first_driver,
            result.second_driver,
            result.third_driver,
            result.fourth_driver,
            result.driver_team,
            result.driver_engine,
            result.gap,
            result.lapped,
            result.penalty,
            result.total,
            points.points
        ])

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало

    # Отправляем сообщение о завершении
    await message.answer_document(
        document=BufferedInputFile(output.read(), filename='results.xlsx')
    )

    output.close()


# Этот хэндлер будет срабатывать на отправку команды /resultxls2
@router.message(Command(commands='resultxls2'), StateFilter(default_state))
async def process_result_xls2_command(message: Message):
    gp = get_actual_gp()
    data = show_result(gp)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[4].height = 22.5
    # Объединяем ячейки в третьем и четвертом столбцах (C и D)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=12)
    ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
    ws['D4'] = f'FORMULA 1 {get_name_gp(gp).upper()} GRAND PRIX'
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = r'logos\Shirokoe_logo_bez_fona_silli.png'  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 7  # % от оригинального размера
    # Рассчитываем новый размер 57,33 -  20
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'C1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Записываем заголовки
    headers = ['POS', '№', 'DRIVER', 'TEAM', None, 'DR1', 'DR2', 'DR3', 'DR4', 'TM', 'ENG', 'DIFF', 'LAP', 'PEN', 'PTS',
               'CH.PTS']
    ws.append(headers)

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.alignment = Alignment(vertical='center')
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=4, end_row=ws.max_row, end_column=5)
    teams_fonts: dict = get_teams_fonts_colors()

    maximus: dict = get_maximus(gp)

    # Записываем данные
    for index, (user, result, points) in enumerate(data, 1):
        ws.append([
            index,
            user.number,
            user.name,
            get_user_team(user.id_telegram),
            None,
            result.first_driver,
            result.second_driver,
            result.third_driver,
            result.fourth_driver,
            result.driver_team,
            result.driver_engine,
            result.gap,
            result.lapped,
            result.penalty,
            result.total,
            points.points
        ])
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
        black_fill = PatternFill(start_color='000001', end_color='000001', fill_type='solid')  # Черный цве
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical='center')
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
                cell.font = wight_font  # Устанавливаем белый шрифт
            else:
                if (cell.column_letter in ['F', 'G'] and cell.value == maximus['max1']) or (
                        cell.column_letter == 'H' and cell.value == maximus['max2']) or (
                        cell.column_letter == 'I' and cell.value == maximus['max3']) or (
                        cell.column_letter in ['L', 'M'] and cell.value == 10):
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=False, color='ED7D31')
                else:
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=False, color='FFFFFF')
            cell.fill = black_fill  # Устанавливаем черный фон

        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(get_user_team(user.id_telegram), None):
            team = teams_fonts.get(get_user_team(user.id_telegram), None)
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            ws.cell(row=ws.max_row, column=2).font = font_number
            ws.cell(row=ws.max_row, column=3).font = font
            ws.cell(row=ws.max_row, column=4).font = font
            ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=3).fill = fill
            ws.cell(row=ws.max_row, column=4).fill = fill
            ws.cell(row=ws.max_row, column=5).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = f'logos\\{team['logo']}'  # Укажите путь к вашему изображению
            else:
                img_path = r'logos\personal.png'  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)
        else:

            img_path = r'logos\personal.png'  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

        # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A'] + ws['B'] + ws['D']:
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['C'].width = 35.7  # Третий столбец
    ws.column_dimensions['D'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['E'].width = 8.7  # Пятый столбец
    ws.column_dimensions[ws.cell(row=7, column=ws.max_column).column_letter].width = 10.7  # Последний столбец

    # Цвета 1, 2, 3 места
    ws.cell(row=8, column=1).fill = PatternFill(start_color='bf9000', end_color='bf9000',
                                                fill_type='solid')
    ws.cell(row=9, column=1).fill = PatternFill(start_color='7c7c7c', end_color='7c7c7c',
                                                fill_type='solid')
    ws.cell(row=10, column=1).fill = PatternFill(start_color='c55a11', end_color='c55a11',
                                                 fill_type='solid')

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало

    # Отправляем сообщение о завершении
    await message.answer_document(
        document=BufferedInputFile(output.read(), filename=f'results {get_name_gp(gp)}.xlsx')
    )

    output.close()


# Этот хэндлер будет срабатывать на отправку команды /calculation
@router.message(Command(commands='calculation'), StateFilter(default_state))
async def process_calculation_command(message: Message):
    gp = get_actual_gp()
    if check_res(gp):
        await message.answer(f'Вы уже сделали расчет для этого GP', isable_web_page_preview=True)
    else:
        results_predict_gp = calculation_drivers(gp)
        # Преобразование словаря в DataFrame
        df = pd.DataFrame(list(results_predict_gp.items()), columns=['Ключ', 'Значение'])

        # Сохраняем книгу в BytesIO
        output = BytesIO()
        output.seek(0)  # Перемещаем указатель в начало
        # Сохранение DataFrame в Excel-файл
        df.to_excel(output, index=False)
        output.seek(0)  # Перемещаем указатель в начало
        await message.answer_document(
            document=BufferedInputFile(output.read(), filename='gp_points.xlsx')
        )
        await message.answer(f'Расчет результатов выполнен')
        output.close()


# Этот хэндлер будет срабатывать на отправку команды /championship
@router.message(Command(commands='championship'), StateFilter(default_state))
async def process_championship_command(message: Message):
    points_all: dict = show_points()
    text_for_answer = f'POS|DRIVER                   |CH.PTS|\n'
    for index, user in enumerate(
            sorted(points_all, key=lambda x: sum([i for i in x.values() if isinstance(i, int)]), reverse=True), 1):
        text_for_answer += f'{index:<3}|{user['User']:<25}|{sum([i for i in user.values() if isinstance(i, int)]):<3}|\n'
    await message.answer(f'<code>{text_for_answer}</code>')


# Этот хэндлер будет срабатывать на отправку команды /championship_full
@router.message(Command(commands='championship_full'), StateFilter(default_state))
async def process_championship_full_command(message: Message):
    points_list: List[dict] = show_points_all(2024)

    for entry in points_list:
        entry['CH.PTS'] = sum(entry[key] for key in entry if key != 'User' and key != 'Team' and entry[key])

    # Сортируем по общему количеству очков
    points_list.sort(key=lambda x: x['CH.PTS'], reverse=True)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Championship Points"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[4].height = 22.5

    # Объединяем ячейки в третьем и четвертом столбцах (C и D)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=12)
    ws.merge_cells(start_row=5, start_column=4, end_row=5, end_column=12)
    ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
    ws['D4'] = f'FORMULA 1 FANTASY SERIES BY SILLY FORMULA'
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=5, column=4).font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000000')
    ws['D5'] = f"DRIVER'S CHAMPIONSHIP"
    ws['D5'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = r'logos\Shirokoe_logo_bez_fona_silli.png'  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 7  # % от оригинального размера
    # Рассчитываем новый размер
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'C1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Заголовки таблицы
    header = ['POS'] + ['№'] + ['Driver'] + ['Team'] + [''] + [key for key in points_list[0] if
                                                               key not in ['User', 'CH.PTS', 'Number', 'Team',
                                                                           'Image']] + ['CH.PTS']
    ws.append(header)  # Добавляем заголовки в первую строку
    ws.row_dimensions[ws.max_row].height = 17

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=4, end_row=ws.max_row, end_column=5)
    teams_fonts: dict = get_teams_fonts_colors()
    # Добавляем данные в файл
    for num, entry in enumerate(points_list, 1):
        row = [num] + [entry['Number']] + [entry['User']] + [entry['Team']] + [''] + [entry[key] for key in header[5:]]
        ws.append(row)  # Добавляем строку с данными
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
        black_fill = PatternFill(start_color='000001', end_color='000001', fill_type='solid')  # Черный цве
        for cell in ws[ws.max_row]:
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
                cell.font = wight_font  # Устанавливаем белый шрифт
            else:
                cell.font = Font(name='Formula1 Display Regular', size=11, bold=False, color='FFFFFF')
            cell.fill = black_fill  # Устанавливаем черный фон

        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(entry['Team'], None):
            team = teams_fonts[entry['Team']]
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            ws.cell(row=ws.max_row, column=2).font = font_number
            ws.cell(row=ws.max_row, column=3).font = font
            ws.cell(row=ws.max_row, column=4).font = font
            ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=3).fill = fill
            ws.cell(row=ws.max_row, column=4).fill = fill
            ws.cell(row=ws.max_row, column=5).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = f'logos\\{team['logo']}'  # Укажите путь к вашему изображению
            else:
                img_path = r'logos\personal.png'  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

    # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='center')

    for cell in ws['A'] + ws['B'] + ws['D']:
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['C'].width = 35.7  # Третий столбец
    ws.column_dimensions['D'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['E'].width = 8.7  # Пятый столбец
    ws.column_dimensions[ws.cell(row=7, column=ws.max_column).column_letter].width = 11.3  # Третий столбец

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало

    # Отправляем сообщение о завершении
    await message.answer_document(
        document=BufferedInputFile(output.read(), filename='championship_points.xlsx')
    )
    output.close()


@router.message(Command(commands='championship_team_full'), StateFilter(default_state))
async def championship_team_full_command(message: Message):
    points_list: List[dict] = show_points_team_all(2024)

    for entry in points_list:
        entry['Total Points'] = sum(entry[key] for key in entry if key != 'Team' and entry[key])

    # Сортируем по общему количеству очков
    points_list.sort(key=lambda x: x['Total Points'], reverse=True)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Championship Points"

    # Заголовки таблицы
    header = ['Team'] + [key for key in points_list[0] if
                         key != 'Total Points' and key != 'Team'] + ['Total Points']
    ws.append(header)  # Добавляем заголовки в первую строку

    # Добавляем данные в файл
    for entry in points_list:
        row = [entry['Team']] + [entry[key] for key in header[1:]]
        ws.append(row)  # Добавляем строку с данными

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало

    # Отправляем сообщение о завершении
    await message.answer_document(
        document=BufferedInputFile(output.read(), filename='championship_team_points.xlsx')
    )
    output.close()

# Хэндлер для текстовых сообщений, которые не попали в другие хэндлеры
@router.callback_query()
async def answer_all(message: Message):
    await message.answer(text=LEXICON_RU['unknown_command'])


# Хэндлер для текстовых сообщений, которые не попали в другие хэндлеры
@router.message()
async def answer_all(message: Message):
    await message.answer(text=LEXICON_RU['unknown_command'])
