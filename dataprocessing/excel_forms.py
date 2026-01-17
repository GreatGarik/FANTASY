from datetime import datetime
import os
from typing import List
import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
from database.database import get_actual_gp_async, show_result, show_points_all, get_user_team, show_points_team_all, get_teams_fonts_colors, get_name_gp, get_maximus, get_all_users, get_predictions_by_gp, get_all_teams_players, get_user_places_by_year, counts_selects

async def entry_list():
    users_list: List[dict] = await get_all_users()

    # Сортируем по общему количеству очков
    users_list.sort(key=lambda x: (x['Number'] if x['Number'] else float('inf'), x['User']))

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "ENTRY LIST"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[3].height = 22.5

    # Объединяем ячейки в третьем и четвертом столбцах (C и D)
    #ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4)
    #ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4)
    ws.cell(row=3, column=3).font = Font(name='Formula1 Display Bold', size=10, bold=True, color='000000')
    ws['C3'] = f'FORMULA 1 FANTASY SERIES BY SILLY FORMULA'
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=4, column=3).font = Font(name='Formula1 Display Bold', size=10, bold=True, color='000000')
    ws['C4'] = f"ENTRY LIST"
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = os.path.join('logos', 'Shirokoe_logo_bez_fona_silli.png')  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 6  # % от оригинального размера
    # Рассчитываем новый размер
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'A1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Заголовки таблицы
    header = ['№'] + ['Driver'] + ['Team'] + ['']

    ws.append(header)  # Добавляем заголовки в первую строку
    ws.row_dimensions[ws.max_row].height = 17

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=3, end_row=ws.max_row, end_column=4)
    teams_fonts: dict = await get_teams_fonts_colors()
    # Добавляем данные в файл
    for num, entry in enumerate(users_list):
        row = [entry['Number'] if entry['Number'] else 'N/A'] + [entry['User']] + [entry['Team']] + ['']
        ws.append(row)  # Добавляем строку с данными
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='000000')  # Черный цвет
        #black_fill = PatternFill(start_color='000001', end_color='000001', fill_type='solid')  # Черный цве
        if num % 2 != 0:
            black_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Белый цвет
        else:
            black_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Белый цвет
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical='center')
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
                cell.font = wight_font  # Устанавливаем белый шрифт
            else:
                cell.font = Font(name='Formula1 Display Regular', size=11, bold=False, color='FFFFFF')
            cell.fill = black_fill  # Устанавливаем черный фон

        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(entry['Team'], None):
            team = teams_fonts[entry['Team']]
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            ws.cell(row=ws.max_row, column=1).font = font_number
            ws.cell(row=ws.max_row, column=2).font = font
            ws.cell(row=ws.max_row, column=3).font = font
            ws.cell(row=ws.max_row, column=1).fill = fill
            ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=3).fill = fill
            ws.cell(row=ws.max_row, column=4).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = os.path.join('logos', team['logo'])  # Укажите путь к вашему изображению
            else:
                img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'D{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)
        else:

            img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'D{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

    # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A'] + ws['B'] + ws['D']:
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['B'].width = 35.7  # Третий столбец
    ws.column_dimensions['C'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['D'].width = 9.0  # Пятый столбец

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало
    return output


async def last_stage():
    gp = await get_actual_gp_async()
    data = await show_result(gp)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[4].height = 22.5
    # Объединяем ячейки в третьем и четвертом столбцах (C и D)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=12)
    ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
    gp_name = await get_name_gp(gp)
    ws['D4'] = f'FORMULA 1 {gp_name.upper()} GRAND PRIX'
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = os.path.join('logos', 'Shirokoe_logo_bez_fona_silli.png')  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 7  # % от оригинального размера
    # Рассчитываем новый размер
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'C1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Записываем заголовки
    headers = ['POS', '№', 'DRIVER', 'TEAM', None, 'DR1', 'DR2', 'DR3', 'DR4', 'TM', 'ENG', 'DL1', 'DL2', 'DL3', 'LAP', 'PEN', 'PTS',
               'CH.PTS']
    ws.append(headers)

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                         right=Side(style='thin', color='FFFFFF'),
                         top=Side(style='thin', color='FFFFFF'))
                         #bottom=Side(style='thin', color='000000')

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=4, end_row=ws.max_row, end_column=5)
    teams_fonts: dict = await get_teams_fonts_colors()

    maximus: dict = await get_maximus(gp)

    # Записываем данные
    for index, (user, result, points) in enumerate(data, 1):
        ws.append([
            index,
            user.number,
            user.name,
            await get_user_team(user.id_telegram),
            None,
            result.first_driver,
            result.second_driver,
            result.third_driver,
            result.fourth_driver,
            result.driver_team,
            result.driver_engine,
            result.select_duel1,
            result.select_duel2,
            result.select_duel3,
            result.lapped,
            result.penalty,
            result.total,
            points.points
        ])
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000001')  # Белый цвет
        if index % 2 != 0:
            black_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Белый цвет
        else:
            black_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Белый цвет
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
                cell.font = wight_font  # Устанавливаем белый шрифт
            else:
                if (cell.column_letter in ['F', 'G'] and cell.value == maximus['max1']) or (
                        cell.column_letter == 'H' and cell.value == maximus['max2']) or (
                        cell.column_letter == 'I' and cell.value == maximus['max3']) or (
                        cell.column_letter in ['L', 'M'] and cell.value == 10):
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='ED7D31')
                else:
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
            cell.fill = black_fill  # Устанавливаем черный фон

        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(await get_user_team(user.id_telegram), None):
            team = teams_fonts.get(await get_user_team(user.id_telegram), None)
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            ws.cell(row=ws.max_row, column=2).font = font_number
           # ws.cell(row=ws.max_row, column=3).font = font
            #ws.cell(row=ws.max_row, column=4).font = font
            ws.cell(row=ws.max_row, column=2).fill = fill
           # ws.cell(row=ws.max_row, column=3).fill = fill
            #ws.cell(row=ws.max_row, column=4).fill = fill
           # ws.cell(row=ws.max_row, column=5).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = os.path.join('logos', team['logo'])  # Укажите путь к вашему изображению
            else:
                img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)
        else:

            img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

        # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='left', vertical='center')


    #for cell in ws['A'] + ws['B'] + ws['D']:
    #    cell.alignment = center_alignment
    for cell in ws['C']:
        if cell.row == 7:
            continue  # Пропускаем 7-й ряд
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['C'].width = 35.7  # Третий столбец
    ws.column_dimensions['D'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['E'].width = 9.2  # Пятый столбец
    ws.column_dimensions[ws.cell(row=7, column=ws.max_column).column_letter].width = 10.7  # Последний столбец

    # Цвета 1, 2, 3 места
    ws.cell(row=8, column=1).fill = PatternFill(start_color='FFC50D', end_color='FFC50D',
                                                fill_type='solid')
    ws.cell(row=9, column=1).fill = PatternFill(start_color='A3A3A3', end_color='A3A3A3',
                                                fill_type='solid')
    ws.cell(row=10, column=1).fill = PatternFill(start_color='BC5610', end_color='BC5610',
                                                 fill_type='solid')

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало
    return output

def sort_points(entry):
    points =  [entry[key] if entry[key] else 0 for key in entry if key not in ['User', 'Team', 'Number', 'CH.PTS'] and not key.startswith('place')]
    places = [-entry[key] if entry[key] else -10000000 for key in entry if key.startswith('place')]
    total_points = sum(points)
    sorted_point  = sorted(points, reverse=True)
    sorted_places = sorted(places, reverse=True)
    # Находим максимальное значение
    max_value = max(points) if points else None

    # Находим первый индекс максимального значения
    first_max_index = points.index(max_value) if max_value in points else -1
    return total_points, sorted_point, -first_max_index, sorted_places

async def process_championship_by_segment():
    points_list: List[dict] = await show_points_all(2025)

    # исходная сортировка (внешняя общая) — сохраняем порядок для стабилизации
    points_list.sort(key=sort_points, reverse=True)

    if not points_list:
        return BytesIO()

    # извлекаем все этапы из ключей словаря в порядке появления
    def extract_stage_keys(example_entry):
        keys = []
        for k in example_entry.keys():
            if k in ('User', 'Team', 'Number', 'Image'):
                continue
            if k.startswith('place'):
                continue
            keys.append(k)
        return keys

    all_stage_keys = extract_stage_keys(points_list[0])
    # сопоставляем индекс этапа (1..N) в порядке появления
    stage_index = {k: i+1 for i, k in enumerate(all_stage_keys)}

    # сегменты по индексам этапов
    segments = [
        ("Seg 1-12", 1, 12),
        ("Seg 13-24", 13, 24),
        ("Seg 1-8", 1, 8),
        ("Seg 9-16", 9, 16),
        ("Seg 17-24", 17, 24),
    ]

    def keys_in_range(start, end):
        return [k for k in all_stage_keys if start <= stage_index[k] <= end]

    wb = Workbook()
    teams_fonts = await get_teams_fonts_colors()

    # Общие стили
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                         right=Side(style='thin', color='FFFFFF'),
                         top=Side(style='thin', color='FFFFFF'))
    data_row_height = 17

    first_sheet = True
    for seg_name, start, end in segments:
        stage_keys = keys_in_range(start, end)
        if not stage_keys:
            continue

        # создаём лист
        if first_sheet:
            ws = wb.active
            ws.title = seg_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=seg_name)

        # Заголовки и лого
        ws.insert_rows(1, amount=5)
        ws.row_dimensions[3].height = 22.5
        last_header_col = 5 + len(stage_keys)
        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=last_header_col)
        ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=last_header_col)
        ws.cell(row=3, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
        ws['D3'] = 'FORMULA 1 FANTASY SERIES BY SILLY FORMULA'
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000000')
        ws['D4'] = f"DRIVER'S CHAMPIONSHIP — {seg_name}"
        ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

        img_path = os.path.join('logos', 'Shirokoe_logo_bez_fona_silli.png')
        if os.path.exists(img_path):
            img = Image(img_path)
            resize_percentage = 7
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))
            img.anchor = 'C1'
            ws.add_image(img)

        # Формируем заголовок таблицы
        header = ['POS', '№', 'Driver', 'Team', ''] + stage_keys + ['CH.PTS']
        ws.append(header)
        ws.row_dimensions[ws.max_row].height = 17
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=ws.max_row, start_column=4, end_row=ws.max_row, end_column=5)

        # Подготовка строк: рассчитываем ch_pts_seg и собираем список, затем сортируем по убыванию
        rows_for_seg = []
        for entry in points_list:
            ch_pts_seg = 0
            values = []
            for k in stage_keys:
                v = entry.get(k)
                if v is None:
                    val = 0
                else:
                    try:
                        val = int(v)
                    except Exception:
                        try:
                            val = int(float(v))
                        except Exception:
                            val = 0
                values.append(val)
                ch_pts_seg += val
            rows_for_seg.append({'entry': entry, 'values': values, 'ch_pts_seg': ch_pts_seg})

        # сортируем по убыванию очков сегмента; для равенства сохраняется исходный порядок points_list
        rows_for_seg.sort(key=lambda x: x['ch_pts_seg'], reverse=True)

        # Запись в лист
        for pos, item in enumerate(rows_for_seg, 1):
            entry = item['entry']
            values = item['values']
            ch_pts_seg = item['ch_pts_seg']
            row = [pos, entry.get('Number', ''), entry.get('User', ''), entry.get('Team', ''), ''] + values + [ch_pts_seg]
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = data_row_height

            # Стили строк
            row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') if pos % 2 != 0 else PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(vertical='center')
                if cell.column_letter in ('A', 'B', 'C', 'D', 'E'):
                    cell.font = Font(name='Formula1 Display Bold', size=11, bold=False, color='000001')
                else:
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = row_fill

            # team styling + logo
            team_name = entry.get('Team')
            if teams_fonts.get(team_name):
                team = teams_fonts[team_name]
                fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'], fill_type='solid')
                font_number = Font(name=team.get('number_font', 'Formula1 Display Bold'), size=14, bold=True, italic=team.get('number_italic', False), color=team.get('number_color', '000000'))
                ws.cell(row=ws.max_row, column=2).font = font_number
                ws.cell(row=ws.max_row, column=2).fill = fill
                logo_name = team.get('logo') or 'personal.png'
                img_path = os.path.join('logos', logo_name)
            else:
                img_path = os.path.join('logos', 'personal.png')

            if os.path.exists(img_path):
                img = Image(img_path)
                rp = 46
                img.width = int(img.width * (rp / 100))
                img.height = int(img.height * (rp / 100))
                img.anchor = f'E{ws.max_row}'
                ws.add_image(img)

        # Выравнивания и ширины колонок
        center_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws['A'] + ws['B'] + ws['D']:
            cell.alignment = center_alignment

        for column in ws.columns:
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = 7.7
        ws.column_dimensions['C'].width = 35.7
        ws.column_dimensions['D'].width = 41.7
        ws.column_dimensions['E'].width = 9.2
        for idx in range(6, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=ws.max_row, column=idx).column_letter].width = 11.3

        if ws.max_row >= 10:
            ws.cell(row=8, column=1).fill = PatternFill(start_color='FFC50D', end_color='FFC50D', fill_type='solid')
            ws.cell(row=9, column=1).fill = PatternFill(start_color='A3A3A3', end_color='A3A3A3', fill_type='solid')
            ws.cell(row=10, column=1).fill = PatternFill(start_color='BC5610', end_color='BC5610', fill_type='solid')

        ws.sheet_view.showGridLines = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

async def process_championship_full():
    points_list: List[dict] = await show_points_all(datetime.now().year)


    # Сортируем по общему количеству очков
    points_list.sort(key=sort_points, reverse=True)

    for entry in points_list:
        entry['CH.PTS'] = sum(
            entry[key] for key in entry if key != 'User' and key != 'Team' and key != 'Number' and (not key.startswith('place')) and entry[key])

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Championship Points"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[3].height = 22.5

    # Объединяем ячейки в третьем и четвертом столбцах (C и D)
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=25)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=25)
    ws.cell(row=3, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
    ws['D3'] = f'FORMULA 1 FANTASY SERIES BY SILLY FORMULA'
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000000')
    ws['D4'] = f"DRIVER'S CHAMPIONSHIP"
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = os.path.join('logos', 'Shirokoe_logo_bez_fona_silli.png')  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 7  # % от оригинального размера
    # Рассчитываем новый размер
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'C1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Заголовки таблицы
    header = ['POS'] + ['№'] + ['Driver'] + ['Team'] + [''] + [key for key in points_list[0] if
                                                               key not in ['User', 'CH.PTS', 'Number', 'Team',
                                                                           'Image'] and not key.startswith('place')] + ['CH.PTS']
    ws.append(header)  # Добавляем заголовки в первую строку
    ws.row_dimensions[ws.max_row].height = 17

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                         right=Side(style='thin', color='FFFFFF'),
                         top=Side(style='thin', color='FFFFFF'))
                         # bottom=Side(style='thin', color='000000'))

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=4, end_row=ws.max_row, end_column=5)
    teams_fonts: dict = await get_teams_fonts_colors()
    # Добавляем данные в файл
    for num, entry in enumerate(points_list, 1):
        row = [num] + [entry['Number']] + [entry['User']] + [entry['Team']] + [''] + [entry[key] for key in header[5:] if not key.startswith('place')]
        ws.append(row)  # Добавляем строку с данными
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='000001')  # Черный цвет
        if num % 2 != 0:
            black_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Белый цвет
        else:
            black_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Белый цвет
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical='center')
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
                cell.font = wight_font  # Устанавливаем белый шрифт
            else:
                cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = black_fill  # Устанавливаем черный фон

        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(entry['Team'], None):
            team = teams_fonts[entry['Team']]
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            ws.cell(row=ws.max_row, column=2).font = font_number
            #ws.cell(row=ws.max_row, column=3).font = font
            #ws.cell(row=ws.max_row, column=4).font = font
            ws.cell(row=ws.max_row, column=2).fill = fill
            #ws.cell(row=ws.max_row, column=3).fill = fill
            #ws.cell(row=ws.max_row, column=4).fill = fill
            #ws.cell(row=ws.max_row, column=5).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = os.path.join('logos', team['logo'])  # Укажите путь к вашему изображению
            else:
                img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)
        else:
            img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'E{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

    # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A'] + ws['B'] + ws['D']:
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['C'].width = 35.7  # Третий столбец
    ws.column_dimensions['D'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['E'].width = 9.2  # Пятый столбец
    ws.column_dimensions[ws.cell(row=7, column=ws.max_column).column_letter].width = 11.3  # Третий столбец

    # Цвета 1, 2, 3 места
    ws.cell(row=8, column=1).fill = PatternFill(start_color='FFC50D', end_color='FFC50D',
                                                fill_type='solid')
    ws.cell(row=9, column=1).fill = PatternFill(start_color='A3A3A3', end_color='A3A3A3',
                                                fill_type='solid')
    ws.cell(row=10, column=1).fill = PatternFill(start_color='BC5610', end_color='BC5610',
                                                 fill_type='solid')

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало
    return output

def sort_points_team(entry):
    total_points = entry['Points']
    sorted_point  = sorted(entry['all_res'], reverse=True)
    # Находим максимальное значение
    max_value = max(sorted_point) if sorted_point else None

    # Находим первый индекс максимального значения
    first_max_index = entry['all_res'].index(max_value) if max_value in entry['all_res'] else -1
    return total_points, sorted_point, -first_max_index


async def statistic_team_excel():
    points_list: List[dict] = await show_points_team_all(2025)

    if not points_list:
        wb = Workbook()
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    stage_keys = [k for k in points_list[0].keys() if k not in ('Team', 'all_res', 'pos')]

    for entry in points_list:
        stage_values = []
        for k in stage_keys:
            v = entry.get(k)
            if isinstance(v, (int, float)) and v:
                stage_values.append(v)
        entry['MAX Stage'] = max(stage_values) if stage_values else 0
        entry['>=200'] = sum(1 for v in stage_values if v >= 200)
        entry['>=150'] = sum(1 for v in stage_values if v >= 150)
        entry['>=100'] = sum(1 for v in stage_values if v >= 100)
        entry['>=50'] = sum(1 for v in stage_values if v >= 50)

    points_list.sort(key=lambda x: x.get('MAX Stage', 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Championship Teams Points"

    # Заголовки: Team, (логотип), MAX Stage, >=200, >=150, >=100, >=50
    header = ['Team', '', 'MAX Stage', '>=200', '>=150', '>=100', '>=50']
    ws.append(header)
    ws.row_dimensions[ws.max_row].height = 17

    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                         right=Side(style='thin', color='FFFFFF'),
                         top=Side(style='thin', color='FFFFFF'))
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    teams_fonts: dict = await get_teams_fonts_colors()

    for entry in points_list:
        row = [
            entry.get('Team', ''),
            '',  # placeholder for logo in column B
            entry.get('MAX Stage', 0),
            entry.get('>=200', 0),
            entry.get('>=150', 0),
            entry.get('>=100', 0),
            entry.get('>=50', 0),
        ]
        ws.append(row)
        ws.row_dimensions[ws.max_row].height = 17

        row_idx = ws.max_row
        # Чередование заливки строк
        if (row_idx - 1) % 2 == 1:
            row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            row_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
            cell.fill = row_fill

        # Применяем стиль команды к колонке Team и вставляем лого в колонку B
        team_name = entry.get('Team', '')
        if teams_fonts.get(team_name):
            team = teams_fonts[team_name]
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team.get('text_color', '000000'))
            fill = PatternFill(start_color=team.get('background_color', 'FFFFFF'),
                               end_color=team.get('background_color', 'FFFFFF'),
                               fill_type='solid')
            ws.cell(row=row_idx, column=1).font = font
            ws.cell(row=row_idx, column=1).fill = fill

            logo_name = team.get('logo') or 'personal.png'
            img_path = os.path.join('logos', logo_name)
            if os.path.exists(img_path):
                img = Image(img_path)
                # масштабирование как в первой версии (0.46)
                img.width = int(img.width * 0.46)
                img.height = int(img.height * 0.46)
                img.anchor = f'B{row_idx}'
                ws.add_image(img)

    # Выравнивание и ширины колонок — вернуть как в первой версии
    center_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws['A'] + ws['B'] + ws['C']:
        cell.alignment = center_alignment

    for column in ws.columns:
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = 7.7
    ws.column_dimensions['A'].width = 41.7  # Team
    ws.column_dimensions['B'].width = 9.2   # Logo
    ws.column_dimensions['C'].width = 11.3  # MAX Stage
    ws.column_dimensions['G'].width = 11.3


    ws.sheet_view.showGridLines = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output




async def championship_team_full():
    points_list: List[dict] = await show_points_team_all(datetime.now().year)

    for entry in points_list:
        entry['Points'] = sum(entry[key] for key in entry if key != 'Team' and key != 'all_res'  and entry[key])

    # Сортируем по общему количеству очков
    #points_list.sort(key=lambda x: x['Points'], reverse=True)
    points_list.sort(key=sort_points_team, reverse=True)

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Championship Teams Points"

    # Вставляем 5 пустых строк в начале
    ws.insert_rows(1, amount=5)
    ws.row_dimensions[4].height = 22.5

    # Объединяем ячейки
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=24)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=24)
    ws.cell(row=3, column=4).font = Font(name='Formula1 Display Bold', size=18, bold=True, color='000000')
    ws['D3'] = f'FORMULA 1 FANTASY SERIES BY SILLY FORMULA'
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=4, column=4).font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000000')
    ws['D4'] = f"TEAM STANDINGS"
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')

    img_path = os.path.join('logos', 'Shirokoe_logo_bez_fona_silli.png')  # Укажите путь к вашему изображению
    img = Image(img_path)
    # Указываем процент изменения размера
    resize_percentage = 7  # % от оригинального размера
    # Рассчитываем новый размер
    img.width = int(img.width * (resize_percentage / 100))
    img.height = int(img.height * (resize_percentage / 100))
    img.anchor = f'B1'  # Устанавливаем позицию изображения
    ws.add_image(img)

    # Заголовки таблицы
    header = ['POS'] + ['Team'] + [''] + [key for key in points_list[0] if
                                          key != 'Points' and key != 'Team' and key != 'all_res'] + ['Points']
    ws.append(header)  # Добавляем заголовки в первую строку
    ws.row_dimensions[ws.max_row].height = 17

    # Устанавливаем шрифт и фон для заголовков
    header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
    header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Красный цвет

    # Создаем черную границу
    thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                         right=Side(style='thin', color='FFFFFF'),
                         top=Side(style='thin', color='FFFFFF'))
                         #bottom=Side(style='thin', color='000000'))

    for cell in ws[7]:  # Перебираем ячейки заголовка
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Объединяем ячейки в третьем и четвертом столбцах (C и D)
        ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=3)
    teams_fonts: dict = await get_teams_fonts_colors()

    # Добавляем данные в файл
    for num, entry in enumerate(points_list, 1):
        row = [num] + [entry['Team']] + [''] + [entry[key] for key in header[1:] if key != '' and key != 'Team' and key != 'all_res']
        ws.append(row)  # Добавляем строку с данными
        ws.row_dimensions[ws.max_row].height = 17
        wight_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')  # Белый цвет
        if num % 2 != 0:
            black_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Белый цвет
        else:
            black_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Белый цвет
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # if cell.column_letter in ['A', 'B', 'C', 'D', 'E']:
            #    cell.font = wight_font  # Устанавливаем белый шрифт

            cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
            cell.fill = black_fill  # Устанавливаем черный фон


        # Устанавливаем фон для ячейки для команд
        if teams_fonts.get(entry['Team'], None):
            team = teams_fonts[entry['Team']]
            # Устанавливаем фон для ячейки, например, в колонке
            font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team['text_color'])
            fill = PatternFill(start_color=team['background_color'], end_color=team['background_color'],
                               fill_type='solid')
            font_number = Font(name=team['number_font'], size=14, bold=True, italic=team['number_italic'],
                               color=team['number_color'])
            # ws.cell(row=ws.max_row, column=2).font = font_number
            ws.cell(row=ws.max_row, column=2).font = font
            # ws.cell(row=ws.max_row, column=4).font = font
            # ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=2).fill = fill
            ws.cell(row=ws.max_row, column=3).fill = fill
            # ws.cell(row=ws.max_row, column=5).fill = fill
            # Вставляем изображение в четвертый столбец (колонка Е)
            if team['logo']:
                img_path = os.path.join('logos', team['logo'])  # Укажите путь к вашему изображению
            else:
                img_path = os.path.join('logos', 'personal.png')  # Укажите путь к вашему изображению
            img = Image(img_path)
            # Указываем процент изменения размера
            resize_percentage = 46  # % от оригинального размера
            # Рассчитываем новый размер
            img.width = int(img.width * (resize_percentage / 100))
            img.height = int(img.height * (resize_percentage / 100))

            img.anchor = f'C{ws.max_row}'  # Устанавливаем позицию изображения
            ws.add_image(img)

        # Устанавливаем выравнивание по центру для нужных колонок
    center_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws['A'] + ws['B'] + ws['C']:
        cell.alignment = center_alignment

    # Устанавливаем ширину столбцов
    for column in ws.columns:
        column_letter = column[0].column_letter  # Получаем букву столбца
        ws.column_dimensions[column_letter].width = 7.7
    # ws.column_dimensions['C'].width = 35.7  # Третий столбец
    ws.column_dimensions['B'].width = 41.7  # Четвертый столбец
    ws.column_dimensions['C'].width = 9.2  # Пятый столбец
    ws.column_dimensions[ws.cell(row=3, column=ws.max_column).column_letter].width = 11.3  # Третий столбец

    # Цвета 1, 2, 3 места
    ws.cell(row=8, column=1).fill = PatternFill(start_color='FFC50D', end_color='FFC50D',
                                                fill_type='solid')
    ws.cell(row=9, column=1).fill = PatternFill(start_color='A3A3A3', end_color='A3A3A3',
                                                fill_type='solid')
    ws.cell(row=10, column=1).fill = PatternFill(start_color='BC5610', end_color='BC5610',
                                                 fill_type='solid')

    # Скрываем сетку
    ws.sheet_view.showGridLines = False

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Перемещаем указатель в начало
    return output

async def process_calculation_command(data):
    # Разделяем данные на два словаря
    total_data = data[0]
    detailed_data = data[1]

    # Создаем DataFrame для общего количества
    total_df = pd.DataFrame(list(total_data.items()), columns=['Участник', 'Total'])

    # Создаем DataFrame для детализированных данных
    detailed_df = pd.DataFrame.from_dict(detailed_data, orient='index').reset_index()
    detailed_df.rename(columns={'index': 'Участник'}, inplace=True)

    # Объединяем два DataFrame по столбцу 'Участник'
    final_df = pd.merge(total_df, detailed_df, on='Участник', how='outer')

    # Определяем порядок строк
    order = []

    # Добавляем участников, не начинающиеся с 'team' и 'engine'
    for key in detailed_data.keys():
        if not key.startswith('team') and not key.startswith('engine'):
            order.append(key)

    # Добавляем пустую строку
    order.append('')

    # Добавляем участников, начинающиеся с 'team'
    for key in detailed_data.keys():
        if key.startswith('team'):
            order.append(key)

    # Добавляем пустую строку
    order.append('')

    # Добавляем участников, начинающиеся с 'engine'
    for key in detailed_data.keys():
        if key.startswith('engine'):
            order.append(key)

    # Добавляем пустую строку
    order.append('')

    # Добавляем MAX1, MAX2, MAX3
    order.extend(['MAX1', 'MAX2', 'MAX3'])

    # Создаем новый DataFrame с заданным порядком
    ordered_df = final_df.set_index('Участник').reindex(order).reset_index()

    # Сохраняем книгу в BytesIO
    output = BytesIO()
    # Создаем Excel-файл
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        ordered_df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Получаем доступ к рабочей книге и листу
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Делаем заголовки жирными
        for cell in worksheet["1:1"]:
            cell.font = cell.font.copy(bold=True)

        # Делаем первый и второй столбцы жирными
        for row in range(2, len(ordered_df) + 2):  # Начинаем с 2, так как 1 - это заголовок
            worksheet.cell(row=row, column=1).font = worksheet.cell(row=row, column=1).font.copy(bold=True)
            worksheet.cell(row=row, column=2).font = worksheet.cell(row=row, column=2).font.copy(bold=True)

        # Устанавливаем ширину столбцов по содержимому
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    output.seek(0)  # Перемещаем указатель в начало
    return output

async def process_all_predicts():
    gp = await get_actual_gp_async()
    data =  await get_predictions_by_gp(gp)
    df = pd.DataFrame(data)

    # Заменяем None на пустую строку для корректного сохранения в Excel
    df.fillna('', inplace=True)
    output = BytesIO()
    # Сохраняем DataFrame в Excel
    df.to_excel(output, index=False)
    # Сохраняем книгу в BytesIO

    # Сохранение DataFrame в Excel-файл
    #df.to_excel(output, index=False)
    output.seek(0)  # Перемещаем указатель в начало
    return output


async def process_all_teams():
    data = await get_all_teams_players()
    # Создаем список для хранения данных
    rows = []

    # Проходим по каждой команде и ее участниками
    for team in data:
        team_name = team['team_name']
        members = team['members']

        # Получаем до трех участников
        member1_name = members[0]['name'] if len(members) > 0 else ''
        member1_number = members[0]['number'] if len(members) > 0 else ''

        member2_name = members[1]['name'] if len(members) > 1 else ''
        member2_number = members[1]['number'] if len(members) > 1 else ''

        member3_name = members[2]['name'] if len(members) > 2 else ''
        member3_number = members[2]['number'] if len(members) > 2 else ''

        # Добавляем строку с данными
        rows.append({
            'Team Name': team_name,
            'Member 1': member1_name,
            'Num1': member1_number,
            'Member 2': member2_name,
            'Num2': member2_number,
            'Member 3': member3_name,
            'Num3': member3_number
        })

    # Создаем DataFrame из списка строк
    df = pd.DataFrame(rows)

    # Заменяем None на пустую строку для корректного сохранения в Excel
    df.fillna('', inplace=True)

    # Создаем Excel файл с помощью openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        # Получаем доступ к рабочей книге и листу
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Устанавливаем ширину столбцов по содержимому
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # Добавляем немного отступа
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    output.seek(0)  # Перемещаем указатель в начало
    return output



async def export_places_summary_by_year():
    rows = await get_user_places_by_year(2025)

    counts = defaultdict(lambda: {
        'first': 0, 'podium': 0, 'top5': 0, 'top10': 0,
        'top20': 0, 'top30': 0, 'top40': 0, 'top50': 0
    })
    user_number = {}
    user_name = {}
    user_team = {}  # uid -> team name (or '')

    for user, place, team_name in rows:
        if getattr(user, 'id', None) is None:
            continue
        uid = user.id
        user_number[uid] = 'N/A' if getattr(user, 'number', None) is None else user.number
        user_name[uid] = getattr(user, 'name', '') or ''
        user_team[uid] = team_name or ''

        if place == 1:
            counts[uid]['first'] += 1
        if 1 <= place <= 3:
            counts[uid]['podium'] += 1
        if 1 <= place <= 5:
            counts[uid]['top5'] += 1
        if 1 <= place <= 10:
            counts[uid]['top10'] += 1
        if 1 <= place <= 20:
            counts[uid]['top20'] += 1
        if 1 <= place <= 30:
            counts[uid]['top30'] += 1
        if 1 <= place <= 40:
            counts[uid]['top40'] += 1
        if 1 <= place <= 50:
            counts[uid]['top50'] += 1

    def build_rows(key, col_name):
        items = [(uid, data[key]) for uid, data in counts.items() if data[key] > 0]
        items.sort(key=lambda x: x[1], reverse=True)
        out = []
        for uid, cnt in items:
            out.append({
                '№': user_number.get(uid, ''),
                'Name': user_name.get(uid, ''),
                'Team': user_team.get(uid, ''),
                '': '',  # пустая колонка-буфер после Team
                col_name: cnt
            })
        return out

    sheets = {
        'WINS': ('first', 'WINS'),
        'PODIUMS': ('podium', 'PODIUMS'),
        'Top-5': ('top5', 'Top-5'),
        'Top-10': ('top10', 'Top-10'),
        'Top-20': ('top20', 'Top-20'),
        'Top-30': ('top30', 'Top-30'),
        'Top-40': ('top40', 'Top-40'),
        'Top-50': ('top50', 'Top-50'),
    }

    teams_fonts = await get_teams_fonts_colors()  # team_name -> {background_color, text_color, number_font, number_italic, number_color, logo}

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Общие стили, используемые в листах
        header_font = Font(name='Formula1 Display Bold', size=11, bold=False, color='FFFFFF')
        header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        thin_border = Border(left=Side(style='thin', color='FFFFFF'),
                             right=Side(style='thin', color='FFFFFF'),
                             top=Side(style='thin', color='FFFFFF'))
        data_row_height = 17  # высота строк данных

        for sheet_name, (key, col_name) in sheets.items():
            df = pd.DataFrame(build_rows(key, col_name), columns=['№', 'Name', 'Team', '', col_name])
            if df.empty:
                df = pd.DataFrame(columns=['№', 'Name', 'Team', '', col_name])
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]

            # Заголовок
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Единая высота строк для данных
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = data_row_height

            # Оформление строк: чередование и применение стиля команды, вставка логотипа в колонку D (4)
            for row_idx in range(2, worksheet.max_row + 1):
                if (row_idx - 2) % 2 == 0:
                    row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                else:
                    row_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if col_idx in (1, 2, 3):
                        cell.font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000001')
                    else:
                        cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
                    cell.fill = row_fill

                team_name = worksheet.cell(row=row_idx, column=3).value or ''
                team_style = teams_fonts.get(team_name)
                if team_style:
                    fill = PatternFill(start_color=team_style['background_color'], end_color=team_style['background_color'], fill_type='solid')
                    text_font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team_style['text_color'])
                    number_font = Font(name=team_style.get('number_font', 'Formula1 Display Bold'),
                                       size=14,
                                       bold=True,
                                       italic=bool(team_style.get('number_italic', False)),
                                       color=team_style.get('number_color', '000000'))
                    worksheet.cell(row=row_idx, column=1).font = number_font
                    worksheet.cell(row=row_idx, column=1).fill = fill
                    worksheet.cell(row=row_idx, column=2).font = text_font
                    worksheet.cell(row=row_idx, column=2).fill = fill
                    worksheet.cell(row=row_idx, column=3).font = text_font
                    worksheet.cell(row=row_idx, column=3).fill = fill

                    # вставка логотипа в колонку D (4)
                    logo_name = team_style.get('logo') or 'personal.png'
                    img_path = os.path.join('logos', logo_name)
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        resize_percentage = 46  # % от оригинального размера
                        img.width = int(img.width * (resize_percentage / 100))
                        img.height = int(img.height * (resize_percentage / 100))
                        img.anchor = f'D{row_idx}'
                        worksheet.add_image(img)
                else:
                    # дефолтный логотип
                    img_path = os.path.join('logos', 'personal.png')
                    if os.path.exists(img_path):
                        img = Image(img_path)
                        resize_percentage = 46  # % от оригинального размера
                        img.width = int(img.width * (resize_percentage / 100))
                        img.height = int(img.height * (resize_percentage / 100))
                        img.anchor = f'D{row_idx}'
                        worksheet.add_image(img)

            # Ширины колонок
            for column in worksheet.columns:
                column_letter = column[0].column_letter
                worksheet.column_dimensions[column_letter].width = 7.7
            worksheet.column_dimensions['B'].width = 35.7  # Name
            worksheet.column_dimensions['C'].width = 41.7  # Team
            worksheet.column_dimensions['D'].width = 9.0   # Logo
            worksheet.column_dimensions['E'].width = 9.0   # Count (First/Podium/Top-5...)

        # --- Дополнительный лист: проценты попаданий в Top-50 ---
        totals = defaultdict(int)
        for user, place, team_name in rows:
            if getattr(user, 'id', None) is None:
                continue
            totals[user.id] += 1

        perc_rows = []
        for uid, total in totals.items():
            top50 = counts.get(uid, {}).get('top50', 0)
            pct = (top50 / total * 100) if total > 0 else 0
            perc_rows.append({
                '№': user_number.get(uid, ''),
                'Name': user_name.get(uid, ''),
                'Team': user_team.get(uid, ''),
                '': '',
                'Total': total,
                'Top-50': top50,
                'Top-50 %': round(pct, 2)
            })

        # сортируем по проценту (по убыванию), затем по total
        perc_rows.sort(key=lambda x: (x['Top-50 %'], x['Total']), reverse=True)

        df_perc = pd.DataFrame(perc_rows, columns=['№', 'Name', 'Team', '', 'Total', 'Top-50', 'Top-50 %'])
        if df_perc.empty:
            df_perc = pd.DataFrame(columns=['№', 'Name', 'Team', '', 'Total', 'Top-50', 'Top-50 %'])
        df_perc.to_excel(writer, sheet_name='PERCENTS', index=False)

        worksheet = writer.sheets['PERCENTS']

        # Заголовок для PERCENTS
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = data_row_height

        for row_idx in range(2, worksheet.max_row + 1):
            if (row_idx - 2) % 2 == 0:
                row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx in (1, 2, 3):
                    cell.font = Font(name='Formula1 Display Bold', size=11, bold=True, color='000001')
                else:
                    cell.font = Font(name='Formula1 Display Regular', size=11, bold=True, color='000001')
                cell.fill = row_fill

            team_name = worksheet.cell(row=row_idx, column=3).value or ''
            team_style = teams_fonts.get(team_name)
            if team_style:
                fill = PatternFill(start_color=team_style['background_color'], end_color=team_style['background_color'], fill_type='solid')
                text_font = Font(name='Formula1 Display Bold', size=11, bold=False, color=team_style['text_color'])
                number_font = Font(name=team_style.get('number_font', 'Formula1 Display Bold'),
                                   size=14,
                                   bold=True,
                                   italic=bool(team_style.get('number_italic', False)),
                                   color=team_style.get('number_color', '000000'))
                worksheet.cell(row=row_idx, column=1).font = number_font
                worksheet.cell(row=row_idx, column=1).fill = fill
                worksheet.cell(row=row_idx, column=2).font = text_font
                worksheet.cell(row=row_idx, column=2).fill = fill
                worksheet.cell(row=row_idx, column=3).font = text_font
                worksheet.cell(row=row_idx, column=3).fill = fill

                logo_name = team_style.get('logo') or 'personal.png'
                img_path = os.path.join('logos', logo_name)
                if os.path.exists(img_path):
                    img = Image(img_path)
                    resize_percentage = 46
                    img.width = int(img.width * (resize_percentage / 100))
                    img.height = int(img.height * (resize_percentage / 100))
                    img.anchor = f'D{row_idx}'
                    worksheet.add_image(img)
            else:
                img_path = os.path.join('logos', 'personal.png')
                if os.path.exists(img_path):
                    img = Image(img_path)
                    resize_percentage = 46
                    img.width = int(img.width * (resize_percentage / 100))
                    img.height = int(img.height * (resize_percentage / 100))
                    img.anchor = f'D{row_idx}'
                    worksheet.add_image(img)

        # ширины колонок для листа PERCENTS
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = 7.7
        worksheet.column_dimensions['B'].width = 35.7
        worksheet.column_dimensions['C'].width = 41.7
        worksheet.column_dimensions['D'].width = 9.0
        # E = Total, F = Top-50, G = Top-50 %
        worksheet.column_dimensions['E'].width = 9.0
        worksheet.column_dimensions['F'].width = 9.0
        worksheet.column_dimensions['G'].width = 12.0

    output.seek(0)
    return output

async def export_counts_to_excel():
    counts = await counts_selects(2025)  # {"total_predicts": int, "drivers": [...], "teams": [...], "engines": [...]}

    rows = []
    # Сводка
    total = counts.get("total_predicts", 0)
    rows.append({"Type": "Summary", "Name": "Total predicts", "Count": total})

    # Сортируем группы по убыванию и добавляем в нужном порядке
    drivers_sorted = sorted(counts.get("drivers", []), key=lambda x: x[1], reverse=True)
    teams_sorted = sorted(counts.get("teams", []), key=lambda x: x[1], reverse=True)
    engines_sorted = sorted(counts.get("engines", []), key=lambda x: x[1], reverse=True)

    # Добавляем пустую строку перед группами для читаемости (опционально)
    rows.append({"Type": "", "Name": "", "Count": ""})

    # Drivers
    rows.append({"Type": "Drivers", "Name": "", "Count": ""})
    for name, cnt in drivers_sorted:
        rows.append({"Type": "Driver", "Name": name, "Count": cnt})

    # пустая строка между группами
    rows.append({"Type": "", "Name": "", "Count": ""})

    # Teams
    rows.append({"Type": "Teams", "Name": "", "Count": ""})
    for name, cnt in teams_sorted:
        rows.append({"Type": "Team", "Name": name, "Count": cnt})

    rows.append({"Type": "", "Name": "", "Count": ""})

    # Engines
    rows.append({"Type": "Engines", "Name": "", "Count": ""})
    for name, cnt in engines_sorted:
        rows.append({"Type": "Engine", "Name": name, "Count": cnt})

    df = pd.DataFrame(rows, columns=["Type", "Name", "Count"])
    df.fillna("", inplace=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]

        # Подгоняем ширину столбцов
        for column_cells in worksheet.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    v = cell.value
                    if v is not None and len(str(v)) > max_length:
                        max_length = len(str(v))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    output.seek(0)
    return output